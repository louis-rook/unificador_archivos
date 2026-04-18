import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def cargar_archivo(ruta):
    """Detecta la extensión y carga el archivo automáticamente."""
    ruta_limpia = str(ruta).strip().lower()
    
    if ruta_limpia.endswith('.xlsx') or ruta_limpia.endswith('.xls'):
        return pd.read_excel(ruta)
    
    try:
        return pd.read_csv(ruta, sep=None, engine='python', encoding='utf-8', on_bad_lines='skip')
    except:
        return pd.read_csv(ruta, sep=None, engine='python', encoding='latin1', on_bad_lines='skip')

def crear_tabla_excel(writer, nombre_hoja, df):
    worksheet = writer.sheets[nombre_hoja]
    
    
    for i, col in enumerate(df.columns):
        if df.empty:
            max_len = len(str(col)) + 2
        else:
            # Calculamos el tamaño máximo sin usar el argumento 'initial'
            longitud_datos = df[col].astype(str).map(len).max()
            max_len = max(longitud_datos, len(str(col))) + 2
            
        worksheet.column_dimensions[get_column_letter(i+1)].width = max_len

    max_row = len(df) + 1
    if max_row == 1: max_row = 2 
    col_letter = get_column_letter(len(df.columns))
    rango = f"A1:{col_letter}{max_row}"

    tabla = Table(displayName=nombre_hoja.replace(" ", "_"), ref=rango)
    estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tabla.tableStyleInfo = estilo
    worksheet.add_table(tabla)

def unificar_documentos():
    root = tk.Tk()
    root.withdraw()

    try:
        messagebox.showinfo("Paso 1 de 2", "Selecciona el reporte de SIESA.")
        archivo_siesa = filedialog.askopenfilename(title="Archivo SIESA", filetypes=[("Archivos", "*.txt *.csv *.xlsx *.xls")])
        if not archivo_siesa: return

        messagebox.showinfo("Paso 2 de 2", "Selecciona el reporte del Cliente a Consolidar.")
        archivo_cliente = filedialog.askopenfilename(title="Archivo cliente", filetypes=[("Archivos", "*.xlsx *.xls *.csv")])
        if not archivo_cliente: return

        # Lectura
        df_siesa = cargar_archivo(archivo_siesa)
        df_cliente = cargar_archivo(archivo_cliente)

        df_siesa.columns = df_siesa.columns.str.strip()
        df_cliente.columns = df_cliente.columns.str.strip()

        clave_siesa, clave_cliente = 'Docto. referencia', 'Referencia'

        if clave_siesa not in df_siesa.columns or clave_cliente not in df_cliente.columns:
            messagebox.showerror("Error", f"Falta la columna '{clave_siesa}' o '{clave_cliente}'.")
            return

        # Limpieza de Siesa (001 y Almacenes cliente)
        df_siesa[clave_siesa] = df_siesa[clave_siesa].fillna('').astype(str).str.strip()
        df_siesa = df_siesa[~df_siesa[clave_siesa].str.upper().str.startswith(('001', 'ALMACENES cliente'))]

        # Renombrar columnas
        df_siesa = df_siesa.rename(columns={col: f"{col} siesa" for col in df_siesa.columns if col != clave_siesa})
        df_cliente = df_cliente.rename(columns={col: f"{col} cliente" for col in df_cliente.columns if col != clave_cliente})

        # Cruce
        df_siesa = df_siesa.rename(columns={clave_siesa: 'Referencia'})
        df_cliente['Referencia'] = df_cliente['Referencia'].fillna('').astype(str).str.strip()
        
        cruce = pd.merge(df_cliente, df_siesa, on='Referencia', how='outer', indicator=True)

        coincidencias = cruce[cruce['_merge'] == 'both'].drop(columns=['_merge']).fillna("")
        no_coincidencias = cruce[cruce['_merge'] != 'both'].drop(columns=['_merge']).fillna("")

        # Guardado
        ruta_descargas = os.path.join(os.environ['USERPROFILE'], 'Downloads')
        archivo_salida = os.path.join(ruta_descargas, "Conciliacion_Siesa_cliente_Final.xlsx")

        with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
            coincidencias.to_excel(writer, sheet_name='Coincidencias', index=False)
            crear_tabla_excel(writer, 'Coincidencias', coincidencias)
            
            no_coincidencias.to_excel(writer, sheet_name='Sin_Coincidencia', index=False)
            if not no_coincidencias.empty:
                crear_tabla_excel(writer, 'Sin_Coincidencia', no_coincidencias)

        messagebox.showinfo("¡Proceso Terminado!", "El cruce fue un cliente.\nEl archivo se abrirá automáticamente.")
        os.startfile(archivo_salida)

    except Exception as e:
        messagebox.showerror("Error", f"Ha ocurrido un problema:\n{e}")

if __name__ == '__main__':
    unificar_documentos()